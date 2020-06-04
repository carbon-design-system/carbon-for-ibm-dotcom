from openpyxl import load_workbook
from selenium import webdriver
import unittest
import time

class calloutQuote(unittest.TestCase):

    def setUp(self):
        global driver
        filepath = "C:/Users/ChetanKanakeri/Desktop/liburl.xlsx"
        wb = load_workbook(filepath)
        sheet = wb.active
        a1 = sheet['A1']
        print("URL  --> ", a1.value)
        driver = webdriver.Chrome()
        driver.get(a1.value)
        driver.maximize_window()
        time.sleep(3)

    def test_calloutquote(self):
        parenthandle = driver.current_window_handle

        CalloutQuote= driver.find_element_by_xpath("//span[contains(text(),'CalloutQuote')]")
        CalloutQuote.click()
        print("Clicked on CallOutQoute Pattern")

       # Click_Default = driver.find_element_by_xpath("//span[contains(text(),'Default')]")
        #Click_Default.click()
        #print("Clicked on default calloutquote")
        time.sleep(2)

        QuoteCopy_Field = driver.find_element_by_xpath("//textarea[@id='Quote (copy): ']")
        QuoteCopy_Field.clear()
        filepath = "C:/Users/ChetanKanakeri/Desktop/liburl.xlsx"
        wb = load_workbook(filepath)
        sheet = wb.active
        a5 = sheet['A5']
        QuoteCopy_Field.send_keys(a5.value)
        print("Entered text in QuoteCopy filed")
        time.sleep(3)

        Full_screen = driver.find_element_by_xpath("//*[contains(@title, 'Open canvas in new tab')]")
        Full_screen.click()
        print("Pattern opened in new tab")
        time.sleep(3)



        handles = driver.window_handles
        for handle in handles:
            if handle not in parenthandle:
                driver.switch_to.window(handle)


                Quote_CopyText = driver.find_element_by_css_selector(".bx--callout__content").value_of_css_property("padding-top")
                print("Pixel diff between top of page and copy text :" + Quote_CopyText)

                Quote_SourceHeading = driver.find_element_by_css_selector(".bx--quote__copy").value_of_css_property("padding-bottom")
                print("Pixel diff between copy text and source heading :" + Quote_SourceHeading)

                CalloutQuote_BottomPage = driver.find_element_by_css_selector(".bx--callout__content").value_of_css_property("padding-bottom")
                print("Pixel value at bottom of page is :" + CalloutQuote_BottomPage)

                Click_Link = driver.find_element_by_xpath("//a[@class='bx--link bx--link-with-icon']")
                Click_Link.click()
                print("Clicked on link")
                time.sleep(3)
                driver.execute_script("window.history.go(-1)")
                time.sleep(3)

                print("Setting browser window to 320px")

                driver.set_window_size(320,315)
                time.sleep(5)
                Quote_CopyText1 = driver.find_element_by_css_selector(".bx--callout__content").value_of_css_property(
                    "padding-top")
                print("Pixel diff between top of page and copy text :" + Quote_CopyText1)

                Quote_SourceHeading1 = driver.find_element_by_css_selector(".bx--quote__copy").value_of_css_property(
                    "padding-bottom")
                print("Pixel diff between copy text and source heading :" + Quote_SourceHeading1)

                CalloutQuote_BottomPage1 = driver.find_element_by_css_selector(".bx--callout__content").value_of_css_property(
                    "padding-bottom")
                print("Pixel value at bottom of page is :" + CalloutQuote_BottomPage1)

                print("Setting browser window to 1056px")

                driver.set_window_size(1058, 515)
                time.sleep(5)

                Quote_CopyText1 = driver.find_element_by_css_selector(".bx--callout__content").value_of_css_property(
                    "padding-top")
                print("Pixel diff between top of page and copy text :" + Quote_CopyText1)

                Quote_SourceHeading1 = driver.find_element_by_css_selector(".bx--quote__copy").value_of_css_property(
                    "padding-bottom")
                print("Pixel diff between copy text and source heading :" + Quote_SourceHeading1)

                CalloutQuote_BottomPage1 = driver.find_element_by_css_selector(
                    ".bx--callout__content").value_of_css_property(
                    "padding-bottom")
                print("Pixel value at bottom of page is :" + CalloutQuote_BottomPage1)



    def tearDown(self):
        driver.quit()
